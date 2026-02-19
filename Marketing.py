import streamlit as st
import pandas as pd
import openpyxl
import random
import os
from groq import Groq
from duckduckgo_search import DDGS

# --- 1. AI CONFIGURATION ---
GROQ_API_KEY = GROQ_API_KEY = st.secrets["GROQ_API_KEY"]
client = Groq(api_key=GROQ_API_KEY)

# --- 2. THE SEARCH ENGINE ---
def get_real_specs(product_name):
    try:
        with DDGS() as ddgs:
            query = f"{product_name} technical specifications features highlights"
            results = [r['body'] for r in ddgs.text(query, max_results=4)]
            return "\n".join(results) if results else "No specific web info found."
    except Exception:
        return "Search temporary unavailable."

# --- 3. EXCEL DATA LOGIC (Specific to 'search' sheet & B4/C4) ---
def load_technodel_search_sheet(path):
    all_items = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        target_sheet_name = next((s for s in wb.sheetnames if s.lower() == "search"), None)
        
        if not target_sheet_name:
            st.error("❌ Could not find a sheet named 'search' in your file.")
            return []

        ws = wb[target_sheet_name]
        # Column B (2) = Name, Column C (3) = Price | Starting at Row 4
        for row in ws.iter_rows(min_row=4, min_col=2, max_col=3):
            name_val = row[0].value 
            price_val = row[1].value 
            if name_val and price_val:
                try:
                    price = int(round(float(str(price_val).replace(',', '')), 0))
                    all_items.append({"name": str(name_val), "price": price})
                except: continue
    except Exception: pass
    return all_items

# --- 4. UI SETTINGS ---
st.set_page_config(page_title="Technodel Marketing Bot 📱", layout="wide")

st.markdown("""
    <style>
    .header-container { display: flex; align-items: center; justify-content: center; gap: 20px; margin-bottom: 30px; }
    .main-title { font-size: 2.5em; font-weight: bold; color: #004a99; margin: 0; }
    .arabic-output { 
        direction: rtl; text-align: right; background-color: #ffffff; 
        padding: 25px; border-radius: 15px; border: 1px solid #eef2f6; 
        font-family: 'Arial'; line-height: 1.8; color: #1a1a1a; font-size: 1.15em;
    }
    .zanouba-header { font-size: 1.4em; color: #d32f2f; font-weight: bold; text-align: center; margin-bottom: 15px; }
    </style>
    """, unsafe_allow_html=True)

# --- 5. LOGO & HEADER ---
LOGO_URL = "https://technodel.net/wp-content/uploads/2024/08/technodel-site-logo-01.webp"
st.markdown(f'''
    <div class="header-container">
        <img src="{LOGO_URL}" width="120">
        <h1 class="main-title">Technodel Marketing Bot</h1>
    </div>
    ''', unsafe_allow_html=True)

# File Selection
files = [f for f in os.listdir('.') if f.endswith(('.xlsx', '.xlsm'))]
selected_file = st.sidebar.selectbox("📂 Select Price List", files) if files else None

col1, col2 = st.columns([1, 2])

with col1:
    if st.button("🎲 Pick Random Product"):
        items = load_technodel_search_sheet(selected_file)
        if items:
            st.session_state.target = random.choice(items)
            st.session_state.output = None
            # Varied Greetings for Zainab
            st.session_state.greeting = random.choice([
                "تفضلي يا أحلى زنوبة، هيدا العرض صار جاهز! ✨",
                "يا هلا بـ أميرة تكنوديل، تفضلي يا زنوبة: 👸",
                "من عيوني يا زنوبة، أحلى بوست لأحلى صبية: 🌷",
                "تكرمي يا زنوبة، شوفي شو حضرنا لليوم: 🔥"
            ])

    if 'target' in st.session_state:
        target = st.session_state.target
        promo_price = int(round(target['price'] * 0.95, 0))
        
        st.success(f"📦 **Selected:** {target['name']}")
        st.info(f"💰 **List Price:** ${target['price']}")
        
        # Dynamic Button Text
        if st.button(f"✨ Generate Offer at {promo_price}$"):
            with st.spinner("Searching specs & writing in Lebanese..."):
                real_info = get_real_specs(target['name'])
                try:
                    completion = client.chat.completions.create(
                        model="llama-3.3-70b-versatile",
                        messages=[
                            {"role": "system", "content": "You are a professional Lebanese tech salesman. You ONLY use Lebanese Arabic dialect (Ammiya). NEVER use formal/Fusha Arabic. Your tone is helpful and high-energy."},
                            {"role": "user", "content": f"""
                                ITEM: {target['name']}
                                ORIGINAL PRICE: ${target['price']}
                                PROMO PRICE: ${promo_price}
                                SPECS: {real_info}
                                
                                Instructions for a True Lebanese Offer:
                                1. Start with a strong Lebanese hook (e.g., 'يا بلاش من تكنوديل', 'شي فاخر من الآخر').
                                2. Describe the item in a detailed paragraph in Lebanese.
                                3. Use the specs found to list 5-6 features in Lebanese bullet points.
                                4. Contrast the price: 'كان بـ ${target['price']} وهلق صار بس بـ ${promo_price}'.
                                5. Mention 1-year warranty ('كفالة سنة') and pickup/delivery in 24h.
                                6. STRICT: No hashtags (#), No links, No formal Arabic words like 'هذا' (use 'هيدا') or 'سوف' (use 'رح').
                            """}
                        ],
                    )
                    st.session_state.output = completion.choices[0].message.content
                except Exception as e:
                    st.error(f"AI Error: {e}")

with col2:
    if st.session_state.get('output'):
        st.markdown(f'<div class="zanouba-header">{st.session_state.greeting}</div>', unsafe_allow_html=True)
        # Display output with line breaks
        formatted_text = st.session_state.output.replace('\n', '<br>')
        st.markdown(f'<div class="arabic-output">{formatted_text}</div>', unsafe_allow_html=True)